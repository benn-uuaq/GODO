import sys
import time
from pathlib import Path
import sqlite3
from openpyxl import Workbook
from datetime import datetime

def get_app_root() -> Path:
    if getattr(sys, "frozen", False):
        base_dir = Path(sys.executable).parent
        internal_dir = base_dir / "_internal"
        if internal_dir.exists() and (internal_dir / "DB").exists():
            return internal_dir
        return base_dir
    else:
        return Path(__file__).resolve().parent.parent

PROJECT_ROOT = get_app_root()

# [추가] DB 폴더가 없을 경우 자동 생성
DB_DIR = PROJECT_ROOT / "DB"
DB_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = DB_DIR / "godo_system.db"
BACKUP_XLSX = DB_DIR / "backup.xlsx"

# =========================
# RobotDB 클래스
# =========================
class RobotDB:
    def __init__(self):
        self.db_path = DB_PATH
        self.backup_xlsx = BACKUP_XLSX

        print("DB_PATH :", self.db_path)
        self._init_db()

    # -------------------------
    # DB 연결
    # -------------------------
    def _get_connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.execute("PRAGMA journal_mode=WAL;")
        return conn

    # -------------------------
    # DB 초기화 및 24개 슬롯 기본 세팅
    # -------------------------
    def _init_db(self):
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS work_goal (
            work_cell TEXT PRIMARY KEY,
            work_count INTEGER DEFAULT 0,
            beaker_type TEXT
        )
        """)
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS work_done (
            work_cell TEXT PRIMARY KEY,
            work_count INTEGER DEFAULT 0
        )
        """)
        
        # no에서 AUTOINCREMENT 제거 (24개 고정)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS result (
            no INTEGER PRIMARY KEY,
            work_cell INTEGER,
            work_point INTEGER,
            beaker_type TEXT,
            qr_value TEXT,
            robot_pose TEXT,
            calib_loc TEXT,
            beaker_loc TEXT,
            loc_deviation TEXT,
            calib_depth TEXT,
            beaker_depth TEXT,
            depth_deviation TEXT,
            work_state TEXT,
            start_time TEXT,
            end_time TEXT
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """)

        # [핵심] result 테이블에 24개의 고정 슬롯이 없으면 생성
        cursor.execute("SELECT COUNT(*) FROM result")
        if cursor.fetchone()[0] == 0:
            for cell in [1, 2]:
                for point in range(1, 13):
                    no = (cell - 1) * 12 + point
                    cursor.execute("""
                        INSERT INTO result (no, work_cell, work_point, work_state) 
                        VALUES (?, ?, ?, 'none')
                    """, (no, cell, point))
                    
        # =====================================================================
        # ★ [추가] 기존 DB 파일에 start_time, end_time 컬럼이 없으면 강제로 밀어 넣기
        # =====================================================================
        try:
            cursor.execute("ALTER TABLE result ADD COLUMN start_time TEXT")
        except sqlite3.OperationalError:
            pass # 이미 컬럼이 존재하면 에러가 나므로 그냥 넘어감 (정상)
            
        try:
            cursor.execute("ALTER TABLE result ADD COLUMN end_time TEXT")
        except sqlite3.OperationalError:
            pass
        # =====================================================================

        conn.commit()
        conn.close()
        self.export_to_excel()

    # -------------------------
    # 엑셀 내보내기 (백업)
    # -------------------------
    def export_to_excel(self):
        conn = self._get_connection()
        cursor = conn.cursor()

        wb = Workbook()
        wb.remove(wb.active) # 기본 시트 제거

        sort_info = {
            "work_goal": "work_cell",
            "work_done": "work_cell",
            "result": "no",
            "settings": "key"
        }

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        db_tables = [row[0] for row in cursor.fetchall() if row[0] != "sqlite_sequence"]

        for table in db_tables:
            sort_col = sort_info.get(table, "rowid") 
            sql = f"SELECT * FROM {table} ORDER BY {sort_col} ASC"

            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
                
                col_names = [desc[0] for desc in cursor.description] if cursor.description else []

                ws = wb.create_sheet(title=table)
                ws.append(col_names)
                for row in rows:
                    ws.append(row)
            except Exception as e:
                print(f"[WARN] {table} 테이블 백업 중 오류 발생: {e}")

        try:
            if not wb.sheetnames:
                wb.create_sheet("Empty")
            wb.save(self.backup_xlsx)
        except PermissionError:
            print("[WARN] 엑셀 파일이 열려있어 백업에 실패했습니다.")
        finally:
            conn.close()

    # =========================================================
    # [Work Management] 작업 준비 및 목표 세팅
    # =========================================================
    def setup_new_work_by_cell(self, work_cell, work_count, beaker_type=""):
        conn = self._get_connection()
        cursor = conn.cursor()
        cell = int(work_cell)
        count = int(work_count)
        
        try:
            # [수정] beaker_type도 함께 저장 및 덮어쓰기
            cursor.execute("""
                INSERT INTO work_goal (work_cell, work_count, beaker_type) 
                VALUES (?, ?, ?)
                ON CONFLICT(work_cell) DO UPDATE SET 
                    work_count=excluded.work_count,
                    beaker_type=excluded.beaker_type
            """, (str(cell), count, str(beaker_type)))
            
            # 2. 해당 셀의 work_done 0으로 초기화
            cursor.execute("""
                INSERT INTO work_done (work_cell, work_count) 
                VALUES (?, 0)
                ON CONFLICT(work_cell) DO UPDATE SET work_count=0
            """, (str(cell),))

            # 3. result 테이블의 해당 셀(12칸) 데이터 초기화 및 상태 분기
            for point in range(1, 13):
                no = (cell - 1) * 12 + point
                
                if point <= count:
                    # 작업 해야 할 포인트: 기존 데이터를 비우고 'waiting' 상태로 만듦
                    cursor.execute("""
                        UPDATE result 
                        SET beaker_type='', qr_value='', robot_pose='', 
                            calib_loc='', beaker_loc='', loc_deviation='', 
                            calib_depth='', beaker_depth='', depth_deviation='', 
                            work_state='waiting'
                        WHERE no=?
                    """, (no,))
                else:
                    # 작업하지 않는 포인트: 모든 값을 '-'로 처리해서 명확히 구분
                    cursor.execute("""
                        UPDATE result 
                        SET beaker_type='-', qr_value='-', robot_pose='-', 
                            calib_loc='-', beaker_loc='-', loc_deviation='-', 
                            calib_depth='-', beaker_depth='-', depth_deviation='-', 
                            work_state='-'
                        WHERE no=?
                    """, (no,))
            
            conn.commit()
            print(f"[DB] 셀 {cell} 새 작업 세팅 완료 (목표 수량: {count}개)")
        except Exception as e:
            print(f"[DB ERROR] 셀 {cell} 새 작업 세팅 실패: {e}")
        finally:
            conn.close()
            self.export_to_excel()

    def update_work_done(self, work_cell, work_count):
        """특정 셀의 작업 완료량을 업데이트합니다."""
        conn = self._get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO work_done (work_cell, work_count)
                VALUES (?, ?)
                ON CONFLICT(work_cell) DO UPDATE SET
                    work_count=excluded.work_count
            """, (str(work_cell), int(work_count)))
            conn.commit()
        except Exception as e:
            print(f"[DB ERROR] work_done 업데이트 실패: {e}")
        finally:
            conn.close()
            self.export_to_excel()

    # =========================================================
    # [Result] 상세 작업 결과 관리 (고정된 24개 행 업데이트)
    # =========================================================
    def update_result_data(self, work_cell, work_point, beaker_type="", qr_value="", 
                           robot_pose="", calib_loc="", beaker_loc="", loc_deviation="", 
                           calib_depth="", beaker_depth="", depth_deviation="", work_state="working",
                           start_time="", end_time=""): # ★ 시간 파라미터 추가
        """해당 셀(지그)과 포인트의 작업 결과를 업데이트합니다."""
        conn = self._get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE result 
                SET beaker_type=?, qr_value=?, robot_pose=?, calib_loc=?, 
                    beaker_loc=?, loc_deviation=?, calib_depth=?, beaker_depth=?, 
                    depth_deviation=?, work_state=?, start_time=?, end_time=?
                WHERE work_cell=? AND work_point=?
            """, (
                str(beaker_type), str(qr_value), str(robot_pose), str(calib_loc), 
                str(beaker_loc), str(loc_deviation), str(calib_depth), str(beaker_depth), 
                str(depth_deviation), str(work_state), str(start_time), str(end_time),
                int(work_cell), int(work_point)
            ))
            conn.commit()
        except Exception as e:
            print(f"[DB ERROR] result 데이터 업데이트 실패: {e}")
        finally:
            conn.close()
            self.export_to_excel()

    def update_result_state(self, work_cell, work_point, new_state):
        """특정 지그의 특정 포인트 상태(work_state)만 변경합니다. (예: working -> done)"""
        conn = self._get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE result 
                SET work_state = ? 
                WHERE work_cell = ? AND work_point = ?
            """, (str(new_state), int(work_cell), int(work_point)))
            conn.commit()
        except Exception as e:
            print(f"[DB ERROR] result 상태 업데이트 실패: {e}")
        finally:
            conn.close()
            self.export_to_excel()

    # =========================================================
    # [Settings] 설정 값 관리
    # =========================================================
    def save_setting(self, key, value):
        conn = self._get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
            conn.commit()
        except Exception as e:
            print(f"[DB] 설정 저장 실패: {e}")
        finally:
            conn.close()
            self.export_to_excel()

    def load_setting(self, key, default_value=None):
        conn = self._get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
            result = cursor.fetchone()
            return result[0] if result else default_value
        except Exception as e:
            print(f"[DB] 설정 로드 실패: {e}")
            return default_value
        finally:
            conn.close()

# # 테스트 실행 예시
# if __name__ == "__main__":
#     TEST = RobotDB()
    
#     TEST.setup_new_work_by_cell(work_cell=1, work_count=8)
#     TEST.setup_new_work_by_cell(work_cell=2, work_count=10)
    
#     # 예시: 셀1의 1번 포인트 작업 결과 저장
#     TEST.update_result_data(work_cell=1, work_point=1, beaker_type="1L", work_state="done")